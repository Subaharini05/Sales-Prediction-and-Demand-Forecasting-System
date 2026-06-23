import os
import io
import datetime
import random
import sqlite3
import mysql.connector
from flask import Flask, request, jsonify, render_template, redirect, url_for, session, send_file
from flask_cors import CORS

from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Import our ML Prediction function
from model.linear_regression import run_sales_prediction

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'coconut_coir_bi_secret_key')
CORS(app)

# Database Configuration (reads environment variables for cloud deployment, falls back to SQLite)
DB_HOST = os.environ.get('DB_HOST', 'localhost')
DB_USER = os.environ.get('DB_USER', 'root')
DB_PASSWORD = os.environ.get('DB_PASSWORD', '')
DB_NAME = os.environ.get('DB_NAME', 'coir_industry_db')

class DatabaseManager:
    def __init__(self):
        self.db_type = 'sqlite'
        self.sqlite_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'coir_database.sqlite')
        self.test_connection()

    def test_connection(self):
        """Test MySQL connection. If fails, fall back to SQLite."""
        # Force SQLite in server environments unless DB_HOST is explicitly configured
        if 'DB_HOST' not in os.environ and 'DATABASE_URL' not in os.environ:
            self.db_type = 'sqlite'
            return

        try:
            # Attempt to connect to MySQL server
            conn = mysql.connector.connect(
                host=DB_HOST,
                user=DB_USER,
                password=DB_PASSWORD,
                connect_timeout=3
            )
            cursor = conn.cursor()
            cursor.execute(f"CREATE DATABASE IF NOT EXISTS {DB_NAME}")
            conn.commit()
            cursor.close()
            conn.close()

            # Now test connecting to the specific database
            conn = mysql.connector.connect(
                host=DB_HOST,
                user=DB_USER,
                password=DB_PASSWORD,
                database=DB_NAME,
                connect_timeout=3
            )
            conn.close()
            self.db_type = 'mysql'
            print("Successfully connected to MySQL database.")
        except Exception as e:
            print(f"MySQL connection failed ({e}). Falling back to SQLite: {self.sqlite_path}")
            self.db_type = 'sqlite'

    def get_connection(self):
        if self.db_type == 'mysql':
            return mysql.connector.connect(
                host=DB_HOST,
                user=DB_USER,
                password=DB_PASSWORD,
                database=DB_NAME
            )
        else:
            conn = sqlite3.connect(self.sqlite_path)
            conn.row_factory = sqlite3.Row
            return conn

    def format_query(self, query):
        """Translate query placeholder %s to ? for SQLite."""
        if self.db_type == 'sqlite':
            return query.replace('%s', '?')
        return query

    def execute_query(self, query, params=None):
        conn = self.get_connection()
        cursor = conn.cursor()
        formatted_query = self.format_query(query)
        try:
            if params:
                cursor.execute(formatted_query, params)
            else:
                cursor.execute(formatted_query)
            
            columns = [col[0] for col in cursor.description] if cursor.description else []
            results = [dict(zip(columns, row)) for row in cursor.fetchall()]
            return results
        except Exception as e:
            print(f"Database error executing query: {query}. Error: {e}")
            raise e
        finally:
            cursor.close()
            conn.close()

    def execute_update(self, query, params=None):
        conn = self.get_connection()
        cursor = conn.cursor()
        formatted_query = self.format_query(query)
        try:
            if params:
                cursor.execute(formatted_query, params)
            else:
                cursor.execute(formatted_query)
            conn.commit()
            last_id = cursor.lastrowid
            return last_id
        except Exception as e:
            print(f"Database error executing update: {query}. Error: {e}")
            conn.rollback()
            raise e
        finally:
            cursor.close()
            conn.close()

    def execute_script(self, script_content):
        conn = self.get_connection()
        try:
            if self.db_type == 'sqlite':
                adapted = script_content
                adapted = adapted.replace("INT AUTO_INCREMENT PRIMARY KEY", "INTEGER PRIMARY KEY AUTOINCREMENT")
                adapted = adapted.replace("DECIMAL(10, 2)", "REAL")
                adapted = adapted.replace("DOUBLE", "REAL")
                conn.executescript(adapted)
            else:
                cursor = conn.cursor()
                statements = script_content.split(';')
                for statement in statements:
                    if statement.strip():
                        cursor.execute(statement)
                conn.commit()
                cursor.close()
        except Exception as e:
            print(f"Script execution failed: {e}")
            raise e
        finally:
            conn.close()

    def initialize_database(self):
        schema_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'database.sql')
        if os.path.exists(schema_path):
            with open(schema_path, 'r') as f:
                schema_content = f.read()
            self.execute_script(schema_content)
            print("Database schema loaded successfully.")
            self.seed_if_empty()
        else:
            print("database.sql not found! Cannot initialize database.")

    def seed_if_empty(self):
        user_check = self.execute_query("SELECT COUNT(*) as cnt FROM Users")
        product_check = self.execute_query("SELECT COUNT(*) as cnt FROM Products")
        
        if user_check[0]['cnt'] == 0:
            print("Seeding default Users...")
            self.execute_update(
                "INSERT INTO Users (Username, Password, Role) VALUES (%s, %s, %s)",
                ("admin", "admin123", "Admin")
            )
            self.execute_update(
                "INSERT INTO Users (Username, Password, Role) VALUES (%s, %s, %s)",
                ("user", "user123", "User")
            )
            
        if product_check[0]['cnt'] == 0:
            print("Seeding default Coir Products...")
            products = [
                ("Coir Mat", "Mats & Flooring Products", 15.00),
                ("Coir Door Mat", "Mats & Flooring Products", 8.50),
                ("Coir Pith Block", "Agricultural & Horticultural Products", 5.00),
                ("Coir Grow Bag", "Agricultural & Horticultural Products", 4.20),
                ("Coir Pot", "Agricultural & Horticultural Products", 2.50),
                ("Coir Mulch Mat", "Agricultural & Horticultural Products", 3.00),
                ("Coir Rope", "Industrial & Fibre Products", 12.00),
                ("Coir Yarn", "Industrial & Fibre Products", 10.00),
                ("Coir Fibre", "Industrial & Fibre Products", 7.50),
                ("Coir Geo Textile", "Industrial & Fibre Products", 18.00)
            ]
            for p in products:
                self.execute_update(
                    "INSERT INTO Products (ProductName, Category, Price) VALUES (%s, %s, %s)",
                    p
                )
            
            # Seed 1 year of daily historical transactions
            sales_check = self.execute_query("SELECT COUNT(*) as cnt FROM Sales")
            if sales_check[0]['cnt'] == 0:
                print("Seeding historical Sales transactions (1 Year)...")
                db_products = self.execute_query("SELECT ProductID, ProductName, Price FROM Products")
                start_date = datetime.date.today() - datetime.timedelta(days=365)
                end_date = datetime.date.today()
                
                curr_date = start_date
                while curr_date <= end_date:
                    num_sales = random.randint(2, 5)
                    selected_products = random.sample(db_products, num_sales)
                    
                    month = curr_date.month
                    for prod in selected_products:
                        qty_base = random.randint(10, 50)
                        season_factor = 1.0
                        if prod['ProductName'] in ["Coir Pith Block", "Coir Grow Bag", "Coir Pot", "Coir Mulch Mat"]:
                            if month in [3, 4, 5]:
                                season_factor = 1.6
                            elif month in [6, 7, 8]:
                                season_factor = 1.3
                        elif prod['ProductName'] in ["Coir Mat", "Coir Door Mat"]:
                            if month in [11, 12, 1]:
                                season_factor = 1.5
                        
                        quantity = int(qty_base * season_factor)
                        price = float(prod['Price'])
                        revenue = quantity * price
                        
                        self.execute_update(
                            "INSERT INTO Sales (Date, ProductID, Quantity, Price, Revenue) VALUES (%s, %s, %s, %s, %s)",
                            (curr_date.strftime("%Y-%m-%d"), prod['ProductID'], quantity, price, revenue)
                        )
                    curr_date += datetime.timedelta(days=1)
                print("Sales transaction logs seeded.")

db = DatabaseManager()

# Initialize tables & seed values on startup
with app.app_context():
    db.initialize_database()

# -------------------------------------------------------------------------
# Security Guard Session Check
# -------------------------------------------------------------------------
@app.before_request
def require_login():
    allowed_routes = ['login_page', 'login_api', 'static']
    if request.endpoint and request.endpoint not in allowed_routes:
        if 'user' not in session:
            return redirect(url_for('login_page'))

# -------------------------------------------------------------------------
# HTML Templates Router Paths
# -------------------------------------------------------------------------
@app.route('/')
def index_page():
    if 'user' in session:
        return redirect(url_for('dashboard_page'))
    return redirect(url_for('login_page'))

@app.route('/login')
def login_page():
    if 'user' in session:
        return redirect(url_for('dashboard_page'))
    return render_template('login.html')

@app.route('/dashboard')
def dashboard_page():
    return render_template('dashboard.html')

@app.route('/products')
def products_page():
    return render_template('products.html')

@app.route('/sales')
def sales_page():
    return render_template('sales.html')

@app.route('/analysis')
def analysis_page():
    return render_template('analysis.html')

@app.route('/prediction')
def prediction_page():
    return render_template('prediction.html')

@app.route('/demand')
def demand_page():
    return render_template('demand.html')

@app.route('/reports')
def reports_page():
    return render_template('reports.html')

# -------------------------------------------------------------------------
# REST APIs: Auth
# -------------------------------------------------------------------------
@app.route('/api/login', methods=['POST'])
def login_api():
    data = request.json or {}
    username = data.get('username')
    password = data.get('password')
    role = data.get('role')
    
    if not username or not password or not role:
        return jsonify({"success": False, "message": "Missing credentials or role."}), 400
        
    user_rows = db.execute_query(
        "SELECT UserID, Username, Role FROM Users WHERE Username = %s AND Password = %s AND Role = %s",
        (username, password, role)
    )
    
    if user_rows:
        user = user_rows[0]
        session['user'] = {
            "id": user['UserID'],
            "username": user['Username'],
            "role": user['Role']
        }
        return jsonify({"success": True, "message": "Login successful.", "user": session['user']})
    return jsonify({"success": False, "message": "Invalid credentials or role."}), 401

@app.route('/api/logout', methods=['POST'])
def logout_api():
    session.pop('user', None)
    return jsonify({"success": True, "message": "Logout successful."})

@app.route('/api/user-session', methods=['GET'])
def user_session_api():
    if 'user' in session:
        return jsonify({"logged_in": True, "user": session['user']})
    return jsonify({"logged_in": False}), 401

# -------------------------------------------------------------------------
# REST APIs: Products CRUD
# -------------------------------------------------------------------------
@app.route('/api/products', methods=['GET'])
def get_products():
    search = request.args.get('search', '')
    category = request.args.get('category', '')
    
    query = "SELECT * FROM Products WHERE 1=1"
    params = []
    
    if search:
        query += " AND ProductName LIKE %s"
        params.append(f"%{search}%")
    if category:
        query += " AND Category = %s"
        params.append(category)
        
    query += " ORDER BY ProductName ASC"
    products = db.execute_query(query, tuple(params))
    return jsonify(products)

@app.route('/api/products', methods=['POST'])
def add_product():
    if session.get('user', {}).get('role') != 'Admin':
        return jsonify({"success": False, "message": "Admin clearance required."}), 403
        
    data = request.json or {}
    name = data.get('ProductName')
    category = data.get('Category')
    price = data.get('Price')
    
    if not name or not category or price is None:
        return jsonify({"success": False, "message": "Missing fields."}), 400
        
    last_id = db.execute_update(
        "INSERT INTO Products (ProductName, Category, Price) VALUES (%s, %s, %s)",
        (name, category, float(price))
    )
    return jsonify({"success": True, "message": "Product created.", "ProductID": last_id})

@app.route('/api/products/<int:pid>', methods=['PUT'])
def edit_product(pid):
    if session.get('user', {}).get('role') != 'Admin':
        return jsonify({"success": False, "message": "Admin clearance required."}), 403
        
    data = request.json or {}
    name = data.get('ProductName')
    category = data.get('Category')
    price = data.get('Price')
    
    db.execute_update(
        "UPDATE Products SET ProductName = %s, Category = %s, Price = %s WHERE ProductID = %s",
        (name, category, float(price), pid)
    )
    return jsonify({"success": True, "message": "Product details updated."})

@app.route('/api/products/<int:pid>', methods=['DELETE'])
def delete_product(pid):
    if session.get('user', {}).get('role') != 'Admin':
        return jsonify({"success": False, "message": "Admin clearance required."}), 403
        
    db.execute_update("DELETE FROM Products WHERE ProductID = %s", (pid,))
    db.execute_update("DELETE FROM Sales WHERE ProductID = %s", (pid,))
    return jsonify({"success": True, "message": "Product and matching sales purged."})

# -------------------------------------------------------------------------
# REST APIs: Sales CRUD
# -------------------------------------------------------------------------
@app.route('/api/sales', methods=['GET'])
def get_sales():
    product_id = request.args.get('product_id', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    
    query = """
        SELECT S.SaleID, S.Date, S.ProductID, P.ProductName, S.Quantity, S.Price, S.Revenue 
        FROM Sales S
        JOIN Products P ON S.ProductID = P.ProductID
        WHERE 1=1
    """
    params = []
    
    if product_id:
        query += " AND S.ProductID = %s"
        params.append(product_id)
    if start_date:
        query += " AND S.Date >= %s"
        params.append(start_date)
    if end_date:
        query += " AND S.Date <= %s"
        params.append(end_date)
        
    query += " ORDER BY S.Date DESC, S.SaleID DESC"
    sales = db.execute_query(query, tuple(params))
    return jsonify(sales)

@app.route('/api/sales', methods=['POST'])
def add_sale():
    data = request.json or {}
    pid = data.get('ProductID')
    date_str = data.get('Date')
    qty = data.get('Quantity')
    
    if not pid or not date_str or qty is None:
        return jsonify({"success": False, "message": "Required fields: ProductID, Date, Quantity."}), 400
        
    prod_row = db.execute_query("SELECT Price FROM Products WHERE ProductID = %s", (pid,))
    if not prod_row:
        return jsonify({"success": False, "message": "Selected product doesn't exist."}), 404
        
    price = float(prod_row[0]['Price'])
    revenue = int(qty) * price
    
    last_id = db.execute_update(
        "INSERT INTO Sales (Date, ProductID, Quantity, Price, Revenue) VALUES (%s, %s, %s, %s, %s)",
        (date_str, pid, int(qty), price, revenue)
    )
    return jsonify({"success": True, "message": "Sale transaction logged.", "SaleID": last_id})

@app.route('/api/sales/<int:sid>', methods=['PUT'])
def edit_sale(sid):
    data = request.json or {}
    pid = data.get('ProductID')
    date_str = data.get('Date')
    qty = data.get('Quantity')
    
    prod_row = db.execute_query("SELECT Price FROM Products WHERE ProductID = %s", (pid,))
    if not prod_row:
        return jsonify({"success": False, "message": "Selected product doesn't exist."}), 404
        
    price = float(prod_row[0]['Price'])
    revenue = int(qty) * price
    
    db.execute_update(
        "UPDATE Sales SET Date = %s, ProductID = %s, Quantity = %s, Price = %s, Revenue = %s WHERE SaleID = %s",
        (date_str, pid, int(qty), price, revenue, sid)
    )
    return jsonify({"success": True, "message": "Transaction record updated."})

@app.route('/api/sales/<int:sid>', methods=['DELETE'])
def delete_sale(sid):
    db.execute_update("DELETE FROM Sales WHERE SaleID = %s", (sid,))
    return jsonify({"success": True, "message": "Sales ledger transaction removed."})

# -------------------------------------------------------------------------
# REST APIs: Analytics & Chart Summaries
# -------------------------------------------------------------------------
@app.route('/api/analytics/kpi', methods=['GET'])
def get_kpis():
    qty_row = db.execute_query("SELECT SUM(Quantity) as qty FROM Sales")
    total_sales_qty = int(qty_row[0]['qty']) if qty_row and qty_row[0]['qty'] is not None else 0
    
    rev_row = db.execute_query("SELECT SUM(Revenue) as rev FROM Sales")
    total_revenue = float(rev_row[0]['rev']) if rev_row and rev_row[0]['rev'] is not None else 0.0
    
    pred_row = db.execute_query("SELECT SUM(PredictedSales) as pred FROM Predictions")
    predicted_sales = float(pred_row[0]['pred']) if pred_row and pred_row[0]['pred'] is not None else 0.0
    
    high_demand_row = db.execute_query("SELECT COUNT(*) as cnt FROM DemandForecast WHERE DemandLevel = 'High'")
    high_demand_count = high_demand_row[0]['cnt'] if high_demand_row else 0
    
    return jsonify({
        "total_sales_qty": total_sales_qty,
        "total_revenue": round(total_revenue, 2),
        "predicted_sales": round(predicted_sales, 2),
        "high_demand_count": high_demand_count
    })

@app.route('/api/analytics/charts', methods=['GET'])
def get_charts():
    if db.db_type == 'sqlite':
        monthly_query = """
            SELECT strftime('%Y-%m', Date) as MonthLabel, SUM(Quantity) as total_qty, SUM(Revenue) as total_rev
            FROM Sales 
            GROUP BY MonthLabel 
            ORDER BY MonthLabel ASC
        """
    else:
        monthly_query = """
            SELECT DATE_FORMAT(Date, '%Y-%m') as MonthLabel, SUM(Quantity) as total_qty, SUM(Revenue) as total_rev
            FROM Sales 
            GROUP BY MonthLabel 
            ORDER BY MonthLabel ASC
        """
    monthly_trends = db.execute_query(monthly_query)
    
    product_sales = db.execute_query("""
        SELECT P.ProductName, SUM(S.Quantity) as total_qty, SUM(S.Revenue) as total_rev
        FROM Sales S
        JOIN Products P ON S.ProductID = P.ProductID
        GROUP BY P.ProductName
        ORDER BY total_rev DESC
    """)
    
    category_sales = db.execute_query("""
        SELECT P.Category, SUM(S.Quantity) as total_qty, SUM(S.Revenue) as total_rev
        FROM Sales S
        JOIN Products P ON S.ProductID = P.ProductID
        GROUP BY P.Category
        ORDER BY total_rev DESC
    """)
    
    return jsonify({
        "monthly_trends": monthly_trends,
        "product_sales": product_sales,
        "category_sales": category_sales
    })

# -------------------------------------------------------------------------
# REST APIs: Machine Learning Predict & Forecast
# -------------------------------------------------------------------------
@app.route('/api/predict', methods=['POST'])
def train_and_predict():
    res = run_sales_prediction(db)
    return jsonify(res)

@app.route('/api/predict/status', methods=['GET'])
def get_predictions_status():
    predictions = db.execute_query("SELECT * FROM Predictions ORDER BY PredictedSales DESC")
    
    if db.db_type == 'sqlite':
        hist_query = """
            SELECT P.ProductName, SUM(S.Quantity) as hist_qty
            FROM Sales S
            JOIN Products P ON S.ProductID = P.ProductID
            WHERE S.Date >= date('now', '-30 days')
            GROUP BY P.ProductName
        """
    else:
        hist_query = """
            SELECT P.ProductName, SUM(S.Quantity) as hist_qty
            FROM Sales S
            JOIN Products P ON S.ProductID = P.ProductID
            WHERE S.Date >= DATE_SUB(CURDATE(), INTERVAL 30 DAY)
            GROUP BY P.ProductName
        """
    hist_sales = db.execute_query(hist_query)
    
    return jsonify({
        "predictions": predictions,
        "historical_comparison": hist_sales
    })

@app.route('/api/forecast', methods=['GET'])
def get_demand_forecast():
    forecasts = db.execute_query("SELECT * FROM DemandForecast ORDER BY ForecastQuantity DESC")
    return jsonify(forecasts)

# -------------------------------------------------------------------------
# Report Exports Core Logic
# -------------------------------------------------------------------------
def get_report_data():
    sales = db.execute_query("""
        SELECT S.SaleID, S.Date, P.ProductName, P.Category, S.Quantity, S.Price, S.Revenue 
        FROM Sales S
        JOIN Products P ON S.ProductID = P.ProductID
        ORDER BY S.Date DESC
    """)
    predictions = db.execute_query("SELECT ProductName, PredictionDate, PredictedSales FROM Predictions ORDER BY PredictedSales DESC")
    forecasts = db.execute_query("SELECT ProductName, ForecastQuantity, DemandLevel FROM DemandForecast ORDER BY ForecastQuantity DESC")
    
    total_sales_qty = db.execute_query("SELECT SUM(Quantity) as qty FROM Sales")[0]['qty'] or 0
    total_revenue = db.execute_query("SELECT SUM(Revenue) as rev FROM Sales")[0]['rev'] or 0.0
    avg_sale = db.execute_query("SELECT AVG(Revenue) as avg_rev FROM Sales")[0]['avg_rev'] or 0.0
    
    top_selling = db.execute_query("""
        SELECT P.ProductName, SUM(S.Quantity) as qty
        FROM Sales S
        JOIN Products P ON S.ProductID = P.ProductID
        GROUP BY P.ProductName
        ORDER BY qty DESC LIMIT 1
    """)
    top_product = top_selling[0]['ProductName'] if top_selling else "N/A"
    
    return {
        "sales": sales,
        "predictions": predictions,
        "forecasts": forecasts,
        "kpis": {
            "total_sales_qty": int(total_sales_qty),
            "total_revenue": round(float(total_revenue), 2),
            "avg_sale": round(float(avg_sale), 2),
            "top_product": top_product
        }
    }

@app.route('/api/report/pdf', methods=['GET'])
def download_pdf_report():
    data = get_report_data()
    buffer = io.BytesIO()
    
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()
    
    primary_color = colors.HexColor("#0B3C5D")
    accent_color = colors.HexColor("#8B5A2B")
    secondary_color = colors.HexColor("#328CC1")
    neutral_dark = colors.HexColor("#1D2731")
    neutral_light = colors.HexColor("#F9F9F9")
    
    title_style = ParagraphStyle('ReportTitle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=22, textColor=primary_color, spaceAfter=6)
    subtitle_style = ParagraphStyle('ReportSubtitle', parent=styles['Normal'], fontName='Helvetica', fontSize=10, textColor=accent_color, spaceAfter=15)
    heading_style = ParagraphStyle('SectionHeading', parent=styles['Heading2'], fontName='Helvetica-Bold', fontSize=14, textColor=primary_color, spaceBefore=12, spaceAfter=8)
    table_header_style = ParagraphStyle('TableHeader', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=9, textColor=colors.white)
    table_cell_style = ParagraphStyle('TableCell', parent=styles['Normal'], fontName='Helvetica', fontSize=9, textColor=neutral_dark)
    
    story = []
    story.append(Paragraph("COCONUT COIR INDUSTRY", title_style))
    story.append(Paragraph(f"Sales Prediction & Demand Forecasting System Report — Generated on {datetime.date.today().strftime('%B %d, %Y')}", subtitle_style))
    story.append(Spacer(1, 10))
    
    kpis = data['kpis']
    kpi_table_data = [
        [Paragraph("<b>Total Sales Quantity:</b>", table_cell_style), Paragraph(f"{kpis['total_sales_qty']:,}", table_cell_style),
         Paragraph("<b>Total Revenue:</b>", table_cell_style), Paragraph(f"${kpis['total_revenue']:,.2f}", table_cell_style)],
        [Paragraph("<b>Average Sale Revenue:</b>", table_cell_style), Paragraph(f"${kpis['avg_sale']:,.2f}", table_cell_style),
         Paragraph("<b>Top Selling Product:</b>", table_cell_style), Paragraph(kpis['top_product'], table_cell_style)]
    ]
    
    kpi_table = Table(kpi_table_data, colWidths=[1.8*inch, 1.8*inch, 1.8*inch, 1.8*inch])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), neutral_light),
        ('PADDING', (0,0), (-1,-1), 8),
        ('BOX', (0,0), (-1,-1), 1.5, primary_color)
    ]))
    
    story.append(Paragraph("Executive Summary KPIs", heading_style))
    story.append(kpi_table)
    story.append(Spacer(1, 15))
    
    story.append(Paragraph("Machine Learning Sales Predictions & Demand Level Classification", heading_style))
    forecast_table_data = [[Paragraph("Product Name", table_header_style), Paragraph("Forecast Quantity", table_header_style), Paragraph("Demand Classification", table_header_style)]]
    
    for f in data['forecasts']:
        level = f['DemandLevel']
        if level == "High":
            lvl_html = f"<font color='{accent_color.hexval()}'><b>High</b></font>"
        elif level == "Medium":
            lvl_html = f"<font color='{secondary_color.hexval()}'><b>Medium</b></font>"
        else:
            lvl_html = "Low"
            
        forecast_table_data.append([
            Paragraph(f['ProductName'], table_cell_style),
            Paragraph(f"{f['ForecastQuantity']:,.2f}", table_cell_style),
            Paragraph(lvl_html, table_cell_style)
        ])
        
    forecast_table = Table(forecast_table_data, colWidths=[3.2*inch, 2.0*inch, 2.0*inch])
    forecast_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), primary_color),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, neutral_light]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E0E0E0")),
        ('PADDING', (0,0), (-1,-1), 6)
    ]))
    story.append(forecast_table)
    doc.build(story)
    
    buffer.seek(0)
    return send_file(
        io.BytesIO(buffer.getvalue()),
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'coir_sales_report_{datetime.date.today().strftime("%Y%m%d")}.pdf'
    )

@app.route('/api/report/excel', methods=['GET'])
def download_excel_report():
    data = get_report_data()
    wb = Workbook()
    
    navy_fill = PatternFill(start_color="0B3C5D", end_color="0B3C5D", fill_type="solid")
    zebra_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
    
    title_font = Font(name="Calibri", size=16, bold=True, color="0B3C5D")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    thin_border = Border(left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'), top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0'))
    
    # Tab 1: Dashboard
    ws_dash = wb.active
    ws_dash.title = "Dashboard Summary"
    ws_dash['A1'] = "COCONUT COIR INDUSTRY - BUSINESS INTELLIGENCE OVERVIEW"
    ws_dash['A1'].font = title_font
    
    ws_dash['A4'] = "KPI Metric"
    ws_dash['B4'] = "Value"
    ws_dash['A4'].font = header_font
    ws_dash['A4'].fill = navy_fill
    ws_dash['B4'].font = header_font
    ws_dash['B4'].fill = navy_fill
    
    kpis = data['kpis']
    kpi_rows = [
        ("Total Sales Quantity", kpis['total_sales_qty']),
        ("Total Revenue", kpis['total_revenue']),
        ("Average Transaction Value", kpis['avg_sale']),
        ("Top Performer Product", kpis['top_product'])
    ]
    
    for idx, (k, v) in enumerate(kpi_rows, start=5):
        ws_dash.cell(row=idx, column=1, value=k).font = bold_font
        cell = ws_dash.cell(row=idx, column=2, value=v)
        cell.font = regular_font
        if isinstance(v, float):
            cell.number_format = '$#,##0.00'
        elif isinstance(v, int):
            cell.number_format = '#,##0'
        ws_dash.cell(row=idx, column=1).border = thin_border
        cell.border = thin_border
        
    ws_dash.column_dimensions['A'].width = 28
    ws_dash.column_dimensions['B'].width = 24
    
    # Tab 2: Forecast
    ws_fore = wb.create_sheet(title="Predictions")
    ws_fore['A1'] = "Sales Predictions & Demand Level Classification"
    ws_fore['A1'].font = title_font
    
    headers = ["Product Name", "Forecast Quantity", "Demand Level"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws_fore.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = navy_fill
        
    for r_idx, f in enumerate(data['forecasts'], start=4):
        ws_fore.cell(row=r_idx, column=1, value=f['ProductName']).font = regular_font
        
        qty_cell = ws_fore.cell(row=r_idx, column=2, value=f['ForecastQuantity'])
        qty_cell.font = regular_font
        qty_cell.number_format = '#,##0.00'
        
        lvl_cell = ws_fore.cell(row=r_idx, column=3, value=f['DemandLevel'])
        lvl_cell.font = bold_font
        
        for c in range(1, 4):
            cell = ws_fore.cell(row=r_idx, column=c)
            cell.border = thin_border
            if r_idx % 2 == 1:
                cell.fill = zebra_fill
                
    ws_fore.column_dimensions['A'].width = 28
    ws_fore.column_dimensions['B'].width = 22
    ws_fore.column_dimensions['C'].width = 18
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(
        io.BytesIO(buffer.getvalue()),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'coir_sales_report_{datetime.date.today().strftime("%Y%m%d")}.xlsx'
    )

if __name__ == '__main__':
    # Binds to dynamic PORT environment variable (highly critical for Render/Heroku hosts)
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=True)
